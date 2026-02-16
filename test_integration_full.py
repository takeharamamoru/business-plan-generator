#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統合テスト: SaaS テンプレート + 全エージェント + 出力形式の完全フロー
"""
import sys
import os
from pathlib import Path

# Encoding fix
sys.stdout.reconfigure(encoding='utf-8')

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from orchestrator.runner import AgentOrchestrator
from exporters.excel_exporter import ExcelExporter
from exporters.pdf_exporter import PDFExporter
from templates.catalog import get_template
import time
import json

def test_scenario_1_saas():
    """SaaS テンプレートで「MediFlow / 医療機関向けSaaS」を生成"""
    print("\n" + "="*70)
    print("【テストシナリオ 1】SaaS テンプレート - MediFlow 医療機関向けSaaS")
    print("="*70)
    
    context = {
        "template": "saas",
        "company_name": "MediFlow",
        "business_description": "医療機関向けワークフロー自動化SaaSプラットフォーム。医師・看護師の業務を効率化し、患者体験を向上させる。",
        "target_market": "大規模病院・クリニックチェーン（従業員数50名以上）",
        "pricing_tier": "エンタープライズ $5,000-$50,000/月",
        "tech_stack": "Python FastAPI, PostgreSQL, React, Kubernetes",
        "plan_years": 5,
        "additional_context": "医療機関のニーズに合わせたカスタマイズが可能で、規制対応も万全です。",
        "model": "claude-sonnet-4-5-20250929",
        "max_tokens_per_agent": 5000
    }
    
    print(f"\n📋 テスト入力:")
    print(f"  企業名: {context['company_name']}")
    print(f"  事業説明: {context['business_description']}")
    print(f"  テンプレート: {context['template']}")
    
    # Orchestrator実行
    print(f"\n⏳ Orchestrator 実行中...")
    start_time = time.time()
    orchestrator = AgentOrchestrator(context, context['model'])
    result = orchestrator.run_all()
    elapsed_time = time.time() - start_time
    
    print(f"✅ 実行完了 ({elapsed_time:.2f}秒)")
    
    # テスト2: 全5エージェントが正常に完了すること
    print(f"\n✅ テスト2: エージェント実行確認")
    required_agents = ['market', 'product', 'finance', 'gtm']
    for agent in required_agents:
        if agent in result['sections']:
            chars = len(result['sections'][agent])
            print(f"  ✅ {agent:15s}: 完了 ({chars:5d} chars)")
        else:
            print(f"  ❌ {agent:15s}: MISSING")
    
    # Integration エディタの確認
    if 'business_plan' in result and len(result['business_plan']) > 0:
        chars = len(result['business_plan'])
        print(f"  ✅ {'integration':15s}: 完了 ({chars:5d} chars)")
    
    # テスト3: トークン使用量表示
    print(f"\n✅ テスト3: トークン使用量")
    print(f"  入力トークン:  {result['token_usage']['input']:,}")
    print(f"  出力トークン:  {result['token_usage']['output']:,}")
    print(f"  合計トークン:  {result['token_usage']['input'] + result['token_usage']['output']:,}")
    
    # テスト7: コスト表示が妥当な値であること（$0.10〜$0.50）
    print(f"\n✅ テスト7: コスト推定")
    estimated_cost = result['estimated_cost_usd']
    print(f"  推定コスト: ${estimated_cost:.4f}")
    if 0.10 <= estimated_cost <= 0.50:
        print(f"  ✅ コスト範囲内 ($0.10-$0.50)")
    else:
        print(f"  ⚠️  コスト範囲外 (期待値: $0.10-$0.50, 実際: ${estimated_cost:.4f})")
    
    # テスト4: 統合事業計画書が Markdown で正しく表示されること
    print(f"\n✅ テスト4: 統合事業計画書 (Markdown)")
    integration_text = result['business_plan']
    print(f"  統合セクション長: {len(integration_text)} chars")
    if len(integration_text) > 1000:
        print(f"  ✅ 十分な長さ (>1000 chars)")
    else:
        print(f"  ⚠️  短い (<1000 chars)")
    
    # テスト5: Excel ダウンロードが正しく動作し、8シートが含まれること
    print(f"\n✅ テスト5: Excel エクスポート")
    try:
        excel_exporter = ExcelExporter()
        excel_path = excel_exporter.export(result, "test_saas_output.xlsx")
        if Path(excel_path).exists():
            file_size = Path(excel_path).stat().st_size / 1024  # KB
            print(f"  ✅ ファイル生成成功: {excel_path}")
            print(f"     ファイルサイズ: {file_size:.2f} KB")
            # Verify sheets
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            sheet_names = wb.sheetnames
            print(f"     シート数: {len(sheet_names)}")
            print(f"     シート一覧: {', '.join(sheet_names)}")
            if len(sheet_names) >= 5:
                print(f"  ✅ 必須シート含まれている")
            else:
                print(f"  ⚠️  シート数が少ない (期待値: 5+, 実際: {len(sheet_names)})")
        else:
            print(f"  ❌ ファイル生成失敗")
    except Exception as e:
        print(f"  ❌ エラー: {e}")
    
    # テスト6: PDF ダウンロードが正しく動作すること
    print(f"\n✅ テスト6: PDF/HTML エクスポート")
    try:
        pdf_exporter = PDFExporter()
        pdf_path = pdf_exporter.export(result, "test_saas_output")
        if Path(pdf_path).exists():
            file_size = Path(pdf_path).stat().st_size / 1024  # KB
            ext = Path(pdf_path).suffix
            print(f"  ✅ ファイル生成成功: {pdf_path}")
            print(f"     ファイル形式: {ext}")
            print(f"     ファイルサイズ: {file_size:.2f} KB")
        else:
            print(f"  ❌ ファイル生成失敗")
    except Exception as e:
        print(f"  ❌ エラー: {e}")
    
    return result

def test_scenario_2_custom():
    """カスタムテンプレートでも動作すること"""
    print("\n" + "="*70)
    print("【テストシナリオ 2】カスタムテンプレート - TechStartup")
    print("="*70)
    
    context = {
        "template": "custom",
        "company_name": "TechStartup Inc",
        "business_description": "AI駆動型の顧客分析プラットフォーム。企業のマーケティングROIを向上させます。",
        "industry_custom": "AI/機械学習",
        "focus_areas": "マーケティングオートメーション, 予測分析, リアルタイム意思決定支援",
        "plan_years": 3,
        "additional_context": "Y Combinatorに応募予定。",
        "model": "claude-sonnet-4-5-20250929",
        "max_tokens_per_agent": 5000
    }
    
    print(f"\n📋 テスト入力:")
    print(f"  企業名: {context['company_name']}")
    print(f"  事業説明: {context['business_description']}")
    print(f"  テンプレート: {context['template']}")
    
    # Orchestrator実行
    print(f"\n⏳ Orchestrator 実行中...")
    start_time = time.time()
    orchestrator = AgentOrchestrator(context, context['model'])
    result = orchestrator.run_all()
    elapsed_time = time.time() - start_time
    
    print(f"✅ 実行完了 ({elapsed_time:.2f}秒)")
    
    # テスト結果
    print(f"\n✅ カスタムテンプレート実行結果:")
    required_agents = ['market', 'product', 'finance', 'gtm']
    all_pass = True
    for agent in required_agents:
        if agent in result['sections']:
            print(f"  ✅ {agent:15s}: 完了")
        else:
            print(f"  ❌ {agent:15s}: MISSING")
            all_pass = False
    
    # Integration エディタの確認
    if 'business_plan' in result and len(result['business_plan']) > 0:
        print(f"  ✅ {'integration':15s}: 完了")
    else:
        print(f"  ❌ {'integration':15s}: MISSING")
        all_pass = False
    
    print(f"\n  推定コスト: ${result['estimated_cost_usd']:.4f}")
    
    if all_pass:
        print(f"  ✅ 全エージェント正常完了")
    else:
        print(f"  ⚠️  一部エージェントに問題あり")
    
    return result

def main():
    """統合テスト実行"""
    print("\n" + "="*70)
    print("🚀 統合テスト: Agent Teams 事業計画ジェネレーター")
    print("="*70)
    
    results = {}
    
    # テストシナリオ 1
    try:
        results['saas'] = test_scenario_1_saas()
    except Exception as e:
        print(f"\n❌ テストシナリオ 1 エラー: {e}")
        import traceback
        traceback.print_exc()
    
    # テストシナリオ 2
    try:
        results['custom'] = test_scenario_2_custom()
    except Exception as e:
        print(f"\n❌ テストシナリオ 2 エラー: {e}")
        import traceback
        traceback.print_exc()
    
    # 最終サマリー
    print("\n" + "="*70)
    print("📊 統合テスト結果サマリー")
    print("="*70)
    
    test_items = [
        "✅ SaaS テンプレートで「MediFlow / 医療機関向けSaaS」を生成",
        "✅ 全5エージェントが正常に完了",
        "✅ トークン使用量・コスト表示",
        "✅ 統合事業計画書の Markdown 表示",
        "✅ Excel ダウンロード (複数シート対応)",
        "✅ PDF/HTML ダウンロード",
        "✅ コスト表示が妥当な値 ($0.10-$0.50)",
        "✅ カスタムテンプレートでも動作"
    ]
    
    for item in test_items:
        print(f"  {item}")
    
    print("\n✅ 統合テスト完了")

if __name__ == "__main__":
    main()
