"""Excel exporter for business plan documents."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelExporter:
    """Export business plan to Excel format."""

    def __init__(self) -> None:
        """Initialize ExcelExporter."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def export(self, result: dict, filename: str = "business_plan.xlsx") -> str:
        """Export business plan to Excel.
        
        Args:
            result: Result dictionary from AgentOrchestrator.run_all()
                   Contains 'business_plan' and 'sections' keys
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        business_plan = result.get("business_plan", "")
        sections = result.get("sections", {})
        
        # Create sheets
        self._create_summary_sheet(business_plan)
        self._create_section_sheet("市場分析", sections.get("market", ""))
        self._create_section_sheet("プロダクト", sections.get("product", ""))
        self._create_section_sheet("財務計画", sections.get("finance", ""))
        self._create_section_sheet("GTM戦略", sections.get("gtm", ""))
        
        # Save to file
        self.workbook.save(filename)
        return filename

    def _create_summary_sheet(self, business_plan: str) -> None:
        """Create summary sheet."""
        ws = self.workbook.create_sheet("サマリー", 0)
        
        ws['A1'] = "事業計画書 - サマリー"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Extract first 30 lines
        lines = business_plan.split('\n')[:30]
        row = 3
        for line in lines:
            if line.strip():
                ws[f'A{row}'] = line[:100]
                row += 1
        
        ws.column_dimensions['A'].width = 80

    def _create_section_sheet(self, sheet_name: str, content: str) -> None:
        """Create sheet for each section."""
        ws = self.workbook.create_sheet(sheet_name)
        
        if not content.strip():
            ws['A1'] = f"{sheet_name}の内容なし"
            return
        
        # Split by lines and write to cells
        lines = content.split('\n')
        for row_idx, line in enumerate(lines[:100], 1):
            if line.strip():
                ws[f'A{row_idx}'] = line[:200]
        
        ws.column_dimensions['A'].width = 100
